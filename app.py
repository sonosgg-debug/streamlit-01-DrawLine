import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import datetime
import io
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from tickers import get_krx_tickers
from screener import run_screener, fit_upper_trendline, screen_single_stock

def fmt_curr(val, ticker):
    if ticker.endswith('.KS') or ticker.endswith('.KQ'):
        return f"{val:,.0f}원"
    else:
        return f"${val:,.2f}"

# 페이지 설정
st.set_page_config(
    page_title="David Ryan's Just Draw the Line Stock Screener",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 커스텀 CSS로 UI 스타일링 (다크 테마 최적화 및 시인성 개선)
st.markdown("""
<style>
    .main-title {
        font-size: 2.2rem;
        font-weight: 700;
        color: #8AB4F8; /* 어두운 바탕에서도 잘 보이고 고급스러운 연파랑색 */
        margin-bottom: 0.2rem;
    }
    .sub-title {
        font-size: 1rem;
        color: #BDC1C6; /* 밝은 회색으로 가독성 향상 */
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #202124; /* 검정색 계열의 배경 적용 */
        color: #F1F3F4; /* 폰트를 밝은 색상으로 강제 지정 */
        padding: 15px;
        border-radius: 8px;
        border-left: 5px solid #8AB4F8; /* 하늘색 테두리 포인트 */
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)


st.markdown('<div class="main-title">David Ryan "Just Draw the Line" 스크리너</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">한국 및 미국 주식시장(KOSPI, KOSDAQ, S&P 500, NASDAQ 100) 종목 중 추세 돌파 및 거래량 동반 종목 발굴 프로그램</div>', unsafe_allow_html=True)


# 기법 소개
with st.expander("ℹ️ 데이비드 라이언의 'Just Draw the Line' 투자 기법이란?"):
    st.markdown("""
    **데이비드 라이언(David Ryan)**은 윌리엄 오닐의 제자이자, 미국 투자 챔피언십 3년 연속 우승에 빛나는 전설적인 투자자입니다.
    
    그는 차트를 지나치게 복잡한 지표(RSI, MACD 등)로 어지럽히지 않고, **오직 주가와 거래량**에 집중하여 선을 그릴 것을 강조했습니다.
    
    ### 📌 핵심 스크리닝 요건
    1. **상승 추세 (Stage 2 Uptrend) 확인**:
       * 주가가 50일, 150일, 200일 이동평균선 위에 위치.
       * 이평선 정배열 (50MA > 150MA > 200MA).
       * 200일 이평선이 최소 1개월 동안 상승 흐름 유지.
       * 주가가 52주 신저가 대비 최소 25% 이상 높고, 52주 신고가 대비 25% 이내에 위치 (박스권 상단 대기).
    2. **하향 추세선 돌파 (Just Draw the Line)**:
       * 최근 하락 조정 기간 동안 고점들을 연결한 상단 저항선(Downtrend line)을 도출.
       * 당일(혹은 직전 영업일) 주가가 이 추세선을 **상향 돌파(Breakout)**하여 마감.
    3. **거래량 확인 (Volume Confirmation)**:
       * 돌파 시점의 거래량이 **최근 20일 평균 거래량 대비 최소 1.5배(150%) 이상** 급증하여 기관의 매수세 확인.
    """)

# 사이드바 설정 영역
st.sidebar.header("⚙️ 스크리닝 조건 설정")

market_choice = st.sidebar.selectbox(
    "대상 시장 선택",
    ["코스피 (KOSPI)", "코스닥 (KOSDAQ)", "전체 시장 (KOSPI + KOSDAQ)", "미국 S&P 500 (US)", "미국 NASDAQ 100 (US)"],
    index=0
)

lookback_period = st.sidebar.slider(
    "추세선 분석 기간 (영업일)",
    min_value=20,
    max_value=90,
    value=40,
    step=5,
    help="최근 고점을 연결하여 추세선을 그릴 분석 윈도우 기간입니다."
)

vol_ratio_thresh = st.sidebar.slider(
    "최소 돌파 거래량 배수",
    min_value=1.0,
    max_value=3.0,
    value=1.5,
    step=0.1,
    help="돌파 당일 거래량이 직전 20일 평균 거래량 대비 몇 배 이상이어야 하는지 결정합니다. (예: 1.5 = 150%)"
)

chunk_size = st.sidebar.number_input(
    "데이터 일괄 요청 크기 (Chunk)",
    min_value=10,
    max_value=100,
    value=50,
    step=10,
    help="yfinance API로 한 번에 다운로드할 종목 개수입니다. 너무 크게 설정하면 API 에러가 발생할 수 있습니다."
)

apply_trend_template = st.sidebar.toggle(
    "장기 상승 추세 조건(Trend Template) 필터",
    value=True,
    help="미너비니의 상승 2단계 정배열 조건을 활성화합니다. 활성화하면 매우 엄격한 상승 추세 종목만 발굴됩니다."
)

breakout_window = st.sidebar.slider(
    "최근 돌파 허용 기간 (영업일)",
    min_value=1,
    max_value=10,
    value=3,
    step=1,
    help="최근 N영업일 이내에 최초 돌파가 일어난 후 추세선 위를 지키고 있는 종목을 허용합니다."
)


# 세션 상태 초기화 (스크리닝 결과 보존용)
if 'screened_df' not in st.session_state:
    st.session_state.screened_df = None
if 'last_run_time' not in st.session_state:
    st.session_state.last_run_time = None
if 'market_type_used' not in st.session_state:
    st.session_state.market_type_used = None

# 스크리닝 시작 버튼
start_screening = st.sidebar.button("🚀 스크리닝 시작", use_container_width=True)

if start_screening:
    market_map = {
        "코스피 (KOSPI)": "KOSPI",
        "코스닥 (KOSDAQ)": "KOSDAQ",
        "전체 시장 (KOSPI + KOSDAQ)": "ALL",
        "미국 S&P 500 (US)": "S&P 500",
        "미국 NASDAQ 100 (US)": "NASDAQ 100"
    }
    selected_market = market_map[market_choice]
    
    with st.spinner("상장 종목 목록을 가져오는 중..."):
        try:
            tickers_df = get_krx_tickers(selected_market)
            total_count = len(tickers_df)
            st.info(f"수집 대상 종목: 총 {total_count}개 (우선주/스팩 필터링 완료)")
        except Exception as e:
            st.error(f"종목 목록 수집 실패: {e}")
            tickers_df = pd.DataFrame()
            
    if not tickers_df.empty:
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # 스크리너 구동을 위한 모니터링 래퍼 함수 (Streamlit progress 연동)
        results = []
        tickers = tickers_df['ticker'].tolist()
        name_map = dict(zip(tickers_df['ticker'], tickers_df['회사명']))
        
        total_tickers = len(tickers)
        chunks = [tickers[i:i + chunk_size] for i in range(0, total_tickers, chunk_size)]
        
        start_time = datetime.datetime.now()

        
        for idx, chunk in enumerate(chunks):
            status_text.text(f"데이터 다운로드 및 분석 중... [{idx+1}/{len(chunks)}] (진행률: {int((idx+1)/len(chunks)*100)}%)")
            progress_bar.progress((idx + 1) / len(chunks))
            
            try:
                # yfinance 멀티 다운로드
                data = yf.download(chunk, period="2y", group_by="ticker", progress=False)

                
                for ticker in chunk:
                    try:
                        if isinstance(data.columns, pd.MultiIndex):
                            ticker_level = 'Ticker' if 'Ticker' in data.columns.names else 1
                            tickers_in_data = data.columns.get_level_values(ticker_level).unique()
                            if ticker not in tickers_in_data:
                                continue
                            df_single = data.xs(ticker, level=ticker_level, axis=1).dropna(subset=['Close', 'High', 'Low', 'Volume'])
                        else:
                            df_single = data.dropna(subset=['Close', 'High', 'Low', 'Volume'])
                            
                        if len(df_single) < 200:
                            continue
                            
                        # 개별 종목 분석
                        res = screen_single_stock(
                            ticker, 
                            name_map[ticker], 
                            df_single, 
                            lookback_period, 
                            vol_ratio_thresh,
                            apply_trend_template=apply_trend_template,
                            breakout_window=breakout_window
                        )
                        if res:
                            results.append(res)

                    except Exception:
                        continue

            except Exception as e:
                pass

                
        # 프로그레스 초기화
        progress_bar.empty()
        status_text.empty()
        
        if results:
            df_final = pd.DataFrame(results)
            # 출력용 한글 칼럼명 매핑
            df_final_display = df_final.rename(columns={
                'ticker': '티커',
                'name': '종목명',
                'price': '현재가',
                'ma50': '50일 MA',
                'ma150': '150일 MA',
                'ma200': '200일 MA',
                'high_52w': '52주 최고가',
                'low_52w': '52주 최저가',
                'vol_ratio': '거래량 비율',
                'breakout_date': '돌파 감지일'
            })
            # 불필요한 칼럼 제거
            df_final_display = df_final_display.drop(columns=['trend_slope', 'trend_intercept'], errors='ignore')
            
            st.session_state.screened_df = df_final_display
            st.session_state.raw_screened_df = df_final # 원본 저장
        else:
            st.session_state.screened_df = pd.DataFrame()
            st.session_state.raw_screened_df = pd.DataFrame()
            
        st.session_state.last_run_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        st.session_state.market_type_used = market_choice

# 결과 디스플레이
if st.session_state.screened_df is not None:
    st.success(f"🔍 스크리닝 완료! (실행 시각: {st.session_state.last_run_time} | 대상: {st.session_state.market_type_used})")
    
    if st.session_state.screened_df.empty:
        st.warning("조건에 부합하는 종목이 발견되지 않았습니다. 분석 기간을 늘리거나 거래량 배수를 낮춰 보세요.")
    else:
        st.markdown(f"#### 포착된 종목 리스트 (총 {len(st.session_state.screened_df)}개)")
        
        # 소수점 포맷팅
        df_format = st.session_state.screened_df.copy()
        
        df_format['현재가'] = df_format.apply(lambda r: fmt_curr(r['현재가'], r['티커']), axis=1)
        df_format['50일 MA'] = df_format.apply(lambda r: fmt_curr(r['50일 MA'], r['티커']), axis=1)
        df_format['150일 MA'] = df_format.apply(lambda r: fmt_curr(r['150일 MA'], r['티커']), axis=1)
        df_format['200일 MA'] = df_format.apply(lambda r: fmt_curr(r['200일 MA'], r['티커']), axis=1)
        df_format['52주 최고가'] = df_format.apply(lambda r: fmt_curr(r['52주 최고가'], r['티커']), axis=1)
        df_format['52주 최저가'] = df_format.apply(lambda r: fmt_curr(r['52주 최저가'], r['티커']), axis=1)
        df_format['거래량 비율'] = df_format['거래량 비율'].map('{:.2f}배'.format)

        
        st.dataframe(df_format, use_container_width=True)

        
        # --- 엑셀 저장 및 다운로드 기능 ---
        st.markdown("### 📥 데이터 익스포트")
        
        # 메모리 버퍼 생성 후 pandas excel 쓰기
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 엑셀 내보내기용 별도 데이터프레임 가공 (raw_screened_df 기반)
            df_excel = st.session_state.raw_screened_df.copy()
            df_excel = df_excel.rename(columns={
                'ticker': '티커',
                'name': '종목명',
                'price': '현재가',
                'ma50': '50일 MA',
                'ma150': '150일 MA',
                'ma200': '200일 MA',
                'high_52w': '52주 최고가',
                'low_52w': '52주 최저가',
                'vol_ratio': '거래량 비율',
                'breakout_date': '돌파 감지일'
            })
            
            # 한국/미국 주식 구분에 따른 소수점 라운딩 및 형변환
            def format_excel_data(row):
                ticker = row['티커']
                is_kr = ticker.endswith('.KS') or ticker.endswith('.KQ')
                
                # 가격 관련 필드들
                price_cols = ['현재가', '50일 MA', '150일 MA', '200일 MA', '52주 최고가', '52주 최저가']
                for col in price_cols:
                    if is_kr:
                        # 한국 주식은 소수점 반올림 후 정수로 변환
                        row[col] = int(round(row[col]))
                    else:
                        # 미국 주식은 소수점 둘째 자리까지 반올림
                        row[col] = round(row[col], 2)
                        
                # 거래량 비율은 공통 소수점 둘째 자리 반올림
                row['거래량 비율'] = round(row['거래량 비율'], 2)
                return row
                
            df_excel = df_excel.apply(format_excel_data, axis=1)
            # 불필요한 분석용 내부 칼럼 제외
            df_excel = df_excel.drop(columns=['trend_slope', 'trend_intercept'], errors='ignore')
            
            # 가공된 데이터프레임을 엑셀에 쓰기
            df_excel.to_excel(writer, index=False, sheet_name='Just Draw the Line')
            
            # openpyxl 객체 제어로 엑셀 서식화
            worksheet = writer.sheets['Just Draw the Line']
            
            # 1. 1행 헤더에 필터/정렬 토글(AutoFilter) 적용
            from openpyxl.utils import get_column_letter
            max_col = worksheet.max_column
            max_row = worksheet.max_row
            if max_row > 0:
                worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
                
            # 2. 열 너비 자동 맞춤 (Auto-fit) 적용
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    # 한글 문자 폭(전각 문자) 보정을 고려한 글자수 산출
                    length = sum(2 if ord(char) > 128 else 1 for char in val)
                    if length > max_len:
                        max_len = length
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # 3. 특정 열 서식 및 정렬 설정 추가 (미국 주식 가격 필드 및 거래량 비율 '0.00' 서식 적용, J열 가운데 정렬)
            from openpyxl.styles import Alignment
            for row_idx in range(2, max_row + 1):
                # A열(티커 - 1번째 열)의 값 분석
                ticker_val = str(worksheet.cell(row=row_idx, column=1).value or '')
                is_kr = ticker_val.endswith('.KS') or ticker_val.endswith('.KQ')
                
                # 미국 주식(달러화 자산)일 경우 가격 열(C~H열, 즉 3~8번째 열)에 소수점 2자리 '0.00' 서식 적용
                if not is_kr:
                    for col_idx in range(3, 9): # 3열(현재가) ~ 8열(52주 최저가)
                        worksheet.cell(row=row_idx, column=col_idx).number_format = '0.00'
                
                # I열 (거래량 비율 - 9번째 열) -> 공통 소수점 2자리 '0.00' 서식 적용
                cell_i = worksheet.cell(row=row_idx, column=9)
                cell_i.number_format = '0.00'
                
                # J열 (돌파 감지일 - 10번째 열) -> 가운데 정렬 적용
                cell_j = worksheet.cell(row=row_idx, column=10)
                cell_j.alignment = Alignment(horizontal='center')


                
        excel_data = output.getvalue()


        
        # 파일명 동적 생성 (JustDrawLine-[MarketCode]-YYYY-MM-DD.xlsx)
        market_code_map = {
            "코스피 (KOSPI)": "KS",
            "코스닥 (KOSDAQ)": "KQ",
            "전체 시장 (KOSPI + KOSDAQ)": "KS&KQ",
            "미국 S&P 500 (US)": "SP",
            "미국 NASDAQ 100 (US)": "NQ"
        }
        market_code = market_code_map.get(st.session_state.market_type_used, "ALL")
        today_str = datetime.date.today().strftime('%Y-%m-%d')
        excel_filename = f"JustDrawLine-{market_code}-{today_str}.xlsx"
        
        st.download_button(
            label="📥 스크리닝 결과 엑셀(.xlsx) 파일 다운로드",
            data=excel_data,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False
        )

        
        # --- 개별 종목 차트 시각화 영역 ---
        st.markdown("---")
        st.markdown("### 📊 종목별 추세선 분석 차트")
        
        # 사용자가 차트로 확인해볼 종목 선택
        selected_stock_name = st.selectbox(
            "시각화할 종목을 선택하세요",
            options=st.session_state.screened_df['종목명'].tolist()
        )
        
        if selected_stock_name:
            # 선택된 종목의 원본 행 데이터 찾기
            row = st.session_state.raw_screened_df[st.session_state.raw_screened_df['name'] == selected_stock_name].iloc[0]
            ticker = row['ticker']
            
            with st.spinner(f"{selected_stock_name} ({ticker}) 주가 데이터 가져오는 중..."):
                # 차트 작성을 위해 2년치 데이터 수집 (여유있게 MA를 그리기 위함)
                df_chart = yf.download(ticker, period="2y", progress=False)
                if isinstance(df_chart.columns, pd.MultiIndex):
                    df_chart.columns = df_chart.columns.droplevel(1)
                df_chart = df_chart.dropna(subset=['Close', 'High', 'Low', 'Volume'])

                
            if not df_chart.empty:
                # 차트용 데이터 계산
                close_prices = df_chart['Close']
                high_prices = df_chart['High']
                low_prices = df_chart['Low']
                volume_prices = df_chart['Volume']
                
                ma50 = close_prices.rolling(window=50).mean()
                ma150 = close_prices.rolling(window=150).mean()
                ma200 = close_prices.rolling(window=200).mean()
                
                # 최근 lookback_period개의 고가로 다시 추세선 피팅
                recent_highs = high_prices.iloc[-lookback_period:].values
                trendline, a, b = fit_upper_trendline(recent_highs)
                
                # 전체 인덱스 중 최근 lookback_period에 해당하는 날짜 목록
                recent_dates = df_chart.index[-lookback_period:]
                
                # Plotly 서브플롯 생성 (주가 캔들스틱 + 거래량 바)
                fig = make_subplots(
                    rows=2, cols=1, 
                    shared_xaxes=True, 
                    vertical_spacing=0.08,
                    row_heights=[0.7, 0.3]
                )
                
                # 1. 캔들스틱 차트 추가
                fig.add_trace(
                    go.Candlestick(
                        x=df_chart.index,
                        open=df_chart['Open'],
                        high=df_chart['High'],
                        low=df_chart['Low'],
                        close=df_chart['Close'],
                        name="주가",
                        increasing_line_color='#EA4335', # 한국 스타일 빨간색 상승
                        decreasing_line_color='#4285F4'  # 한국 스타일 파란색 하락
                    ),
                    row=1, col=1
                )
                
                # 2. 이동평균선 추가
                fig.add_trace(
                    go.Scatter(x=df_chart.index, y=ma50, line=dict(color='#FBBC05', width=1.5), name="50일 MA"),
                    row=1, col=1
                )
                fig.add_trace(
                    go.Scatter(x=df_chart.index, y=ma150, line=dict(color='#34A853', width=1.5), name="150일 MA"),
                    row=1, col=1
                )
                fig.add_trace(
                    go.Scatter(x=df_chart.index, y=ma200, line=dict(color='#EA4335', width=2), name="200일 MA"),
                    row=1, col=1
                )
                
                # 3. "Just Draw the Line" 하향 추세선 오버레이
                if trendline is not None:
                    fig.add_trace(
                        go.Scatter(
                            x=recent_dates,
                            y=trendline,
                            line=dict(color='#FFFFFF', width=3, dash='dash'),
                            name="하향 추세선 (Downtrend Resistance Line)"
                        ),
                        row=1, col=1
                    )
                    
                    # 돌파 지점 하이라이트 (돌파 감지일 기준 동적 위치 탐색)
                    breakout_date_dt = pd.to_datetime(row['breakout_date'])
                    try:
                        breakout_idx = df_chart.index.get_loc(breakout_date_dt)
                        breakout_price = df_chart['Close'].iloc[breakout_idx]
                    except Exception:
                        # 매칭 실패 시 차트 마지막 일자로 대체
                        breakout_idx = -1
                        breakout_date_dt = df_chart.index[-1]
                        breakout_price = df_chart['Close'].iloc[-1]

                    
                    fig.add_annotation(
                        x=breakout_date_dt,
                        y=breakout_price,
                        text="★ 돌파 (Breakout)",
                        showarrow=True,
                        arrowhead=2,
                        arrowsize=1,
                        arrowwidth=2,
                        arrowcolor="#EA4335",
                        ax=0,
                        ay=-40,
                        font=dict(color="#EA4335", size=12, family="Malgun Gothic"),
                        row=1, col=1
                    )
                
                # 4. 거래량 차트 추가
                # 상승/하락일에 따른 거래량 색상 구분
                colors = ['#EA4335' if df_chart['Close'].iloc[i] >= df_chart['Open'].iloc[i] else '#4285F4' for i in range(len(df_chart))]
                fig.add_trace(
                    go.Bar(
                        x=df_chart.index,
                        y=df_chart['Volume'],
                        marker_color=colors,
                        name="거래량"
                    ),
                    row=2, col=1
                )
                
                # 거래량 20일 이동평균 추가
                vol_ma20 = df_chart['Volume'].rolling(window=20).mean()
                fig.add_trace(
                    go.Scatter(
                        x=df_chart.index,
                        y=vol_ma20,
                        line=dict(color='#5F6368', width=1.5),
                        name="20일 거래량 MA"
                    ),
                    row=2, col=1
                )
                
                # 화폐 단위 동적 결정
                is_us_stock = not (ticker.endswith('.KS') or ticker.endswith('.KQ'))
                currency_symbol = '$' if is_us_stock else '원'
                tick_format = ',.2f' if is_us_stock else ',.0f'
                
                # 레이아웃 정밀화
                fig.update_layout(
                    title=f"📈 {selected_stock_name} ({ticker}) 'Just Draw the Line' 분석 차트",
                    yaxis_title=f"주가 ({currency_symbol})",
                    yaxis2_title="거래량 (주)",
                    xaxis_rangeslider_visible=False,
                    height=700,
                    margin=dict(l=50, r=50, t=80, b=50),
                    showlegend=True,
                    legend=dict(
                        orientation="h",
                        y=1.08,
                        xanchor="right",
                        x=1
                    ),

                    hovermode="x unified"
                )
                
                fig.update_yaxes(tickformat=tick_format, row=1, col=1)
                fig.update_yaxes(tickformat=",.0f", row=2, col=1)
                
                # 주말 공백 제거 (주식 시장 휴장일 제외하여 캔들스틱 간격 유지)
                # KOSPI/KOSDAQ은 평일에만 열리므로 날짜 축에서 주말을 제외
                fig.update_xaxes(
                    rangebreaks=[
                        dict(bounds=["sat", "mon"]), # 토요일부터 월요일 아침까지 비활성화
                    ]
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # 추가 설명 카드
                st.markdown(f"""
                <div class="metric-card">
                    <h4>💡 {selected_stock_name} 상세 분석 정보</h4>
                    <ul>
                        <li><b>돌파 발생일:</b> {row['breakout_date']}</li>
                        <li><b>돌파 시점 거래량 폭증 비율:</b> <span style="color:#EA4335; font-weight:bold;">{row['vol_ratio']:.2f}배</span> (이전 20일 평균 거래량 대비)</li>
                        <li><b>52주 최고가 대비 가격:</b> {row['price'] / row['high_52w'] * 100:.1f}% 수준 (최고가: {fmt_curr(row['high_52w'], ticker)})</li>
                        <li><b>52주 최저가 대비 가격:</b> +{ (row['price'] / row['low_52w'] - 1) * 100:.1f}% 상승 상태 (최저가: {fmt_curr(row['low_52w'], ticker)})</li>
                    </ul>
                </div>
                """, unsafe_allow_html=True)

else:
    # 프로그램 최초 진입 시 메인 화면
    st.info("👈 왼쪽 사이드바에서 대상 시장 및 파라미터를 설정한 후 '스크리닝 시작' 버튼을 눌러주세요.")
